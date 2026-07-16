from django.db.models import Sum, Count, Avg, F, Q, DecimalField
from django.db.models.functions import TruncHour, TruncDay, TruncMonth
from django.utils import timezone
from datetime import timedelta, datetime
from decimal import Decimal
from django.db import transaction
from typing import List, Dict, Tuple, Any, Optional
import os
from .models import Product, Category, Unit, Barcode
from organization.models import Warehouse, Branch
from stock.services import add_stock_to_warehouse
from stock.models import Supplier, StockEntry, StockAdjustment
import uuid
from io import BytesIO


def _make_serializable(data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Convert a dictionary to a JSON-serializable format.
    Converts model instances to their IDs.
    """
    serializable_data = {}
    for key, value in data.items():
        if hasattr(value, 'pk'):  # It's a model instance
            serializable_data[key] = value.pk
        elif hasattr(value, 'id') and not isinstance(value, (str, int, float, bool, type(None))):  # Model instance with id
            serializable_data[key] = value.id
        elif isinstance(value, Decimal):
            serializable_data[key] = str(value)
        else:
            serializable_data[key] = value
    return serializable_data


def _generate_unique_sku_from_name(name: str) -> str:
    """Generate a simple unique SKU from product name."""
    base = ''.join(ch for ch in name.upper() if ch.isalnum())[:8] or 'SKU'
    suffix = str(uuid.uuid4())[:6].upper()
    sku = f"{base}-{suffix}"
    attempts = 0
    while Product.objects.filter(sku=sku).exists() and attempts < 5:
        suffix = str(uuid.uuid4())[:6].upper()
        sku = f"{base}-{suffix}"
        attempts += 1
    return sku


def _safe_decimal(value) -> Optional[Decimal]:
    if value is None or value == '':
        return None
    try:
        return Decimal(str(value))
    except Exception:
        return None


def import_products_from_stock_sheet(file_obj, created_by=None, notes: str = '') -> Dict[str, Any]:
    """
    Import products (and suppliers) from an Excel (.xlsx) file using stock sheet columns:
        product name, description, selling prices, cost per unit, total quantity, supplier name, supplier email address
    
    Product matching and creation rules:
        - Products can have the same name and/or description (multiple products allowed)
        - Match by BOTH name AND description (case-insensitive) to determine if product exists
        - If a product with the exact same name AND description exists: reuse it and increment stock quantity
        - If no exact match (both name and description): create a new product
        - Stock quantity is incremented for existing products, new stock entries created for new products
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError('openpyxl is required to import Excel files.') from exc

    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active

    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip().lower() if cell.value is not None else '')

    # Allow flexible header aliases (case-insensitive)
    header_aliases = {
        'product_name': ['product name', 'name', 'product'],
        'description': ['description', 'desc'],
        'selling_price': ['selling prices', 'selling price', 'sell price', 'selling', 'price (selling)'],
        'purchase_price': ['cost per unit', 'purchase price', 'buy price', 'cost', 'price', 'unit cost'],
        'quantity': ['total quantity', 'quantity', 'qty', 'units'],
        'supplier_name': ['supplier name', 'supplier'],
        'supplier_email': ['supplier email address', 'supplier email', 'supplier_email'],
    }

    header_to_key = {}
    for idx, header in enumerate(headers):
        normalized = header.strip().lower() if header else ''
        for key, aliases in header_aliases.items():
            aliases_normalized = [a.strip().lower() for a in aliases]
            if normalized in aliases_normalized:
                header_to_key[idx] = key
                break

    required_keys = set(header_aliases.keys())
    found_keys = set(header_to_key.values())
    missing_keys = required_keys - found_keys
    if missing_keys:
        # Provide user-friendly names for missing columns
        pretty_missing = [alias_list[0].title() for k, alias_list in header_aliases.items() if k in missing_keys]
        raise ValueError(f'Missing required columns: {", ".join(pretty_missing)}')

    created_products = []
    main_wh = None  # cache main warehouse lookup
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        row_data = {header_to_key[i]: row[i].value for i in header_to_key if i < len(row)}
        if all(val in (None, '') for val in row_data.values()):
            continue

        try:
            product_name = (row_data.get('product_name') or '').strip()
            description = (row_data.get('description') or '').strip()
            supplier_name = (row_data.get('supplier_name') or '').strip()
            supplier_email = (row_data.get('supplier_email') or '').strip()
            selling_price = _safe_decimal(row_data.get('selling_price'))
            purchase_price = _safe_decimal(row_data.get('purchase_price'))
            quantity = _safe_decimal(row_data.get('quantity'))

            if not (product_name or description):
                raise ValueError('Missing product name/description.')

            # Supplier lookup or create
            supplier = None
            if supplier_email:
                supplier = Supplier.objects.filter(email__iexact=supplier_email).first()
            if not supplier and supplier_name:
                supplier = Supplier.objects.filter(name__iexact=supplier_name).first()
            if not supplier and supplier_name:
                supplier = Supplier.objects.create(name=supplier_name, email=supplier_email)

            # Product lookup or create with rules:
            # - Match by BOTH name AND description (case-insensitive)
            # - If both name and description match an existing product, reuse it (will increment stock)
            # - Otherwise, create a new product (allows products with same name/description to exist separately)
            product = None
            if product_name and description:
                # Match by both name and description
                product = Product.objects.filter(
                    name__iexact=product_name,
                    description__iexact=description
                ).first()
            elif product_name:
                # Only name provided - match by name and empty/null description
                product = Product.objects.filter(
                    Q(description__isnull=True) | Q(description__exact=''),
                    name__iexact=product_name
                    
                ).first()
            elif description:
                # Only description provided - match by description and empty/null name
                product = Product.objects.filter(
                    Q(name__isnull=True) | Q(name__exact=''),
                    description__iexact=description
                    
                ).first()
            
            # If no matching product found, create a new one
            if not product:
                sku = _generate_unique_sku_from_name(product_name or description or 'SKU')
                product = Product.objects.create(
                    sku=sku,
                    name=product_name or description or sku,
                    description=description
                )

            stock_created = False
            if quantity is not None and purchase_price is not None:
                if main_wh is None:
                    try:
                        main_wh = get_main_warehouse()
                    except Warehouse.DoesNotExist:
                        raise ValueError('No main warehouse found. Please set a warehouse as is_main=True.')

                # Ensure selling price is set on the stock entry; default to purchase price when missing
                selling_price_final = selling_price if selling_price is not None else purchase_price
                if selling_price_final is None:
                    selling_price_final = purchase_price
                selling_price_final = Decimal(str(selling_price_final))

                # If stock entry exists for this product in main warehouse, increment quantity and update prices
                existing_entry = StockEntry.objects.filter(
                    product=product,
                    warehouse=main_wh
                ).order_by('-received_date', '-created_at').first()

                if existing_entry:
                    existing_entry.quantity += quantity
                    existing_entry.purchase_price = purchase_price
                    existing_entry.selling_price = selling_price_final
                    existing_entry.save(update_fields=['quantity', 'purchase_price', 'selling_price'])
                    # Always create a stock adjustment record to keep history
                    StockAdjustment.objects.create(
                        product=product,
                        warehouse=main_wh,
                        adjustment_type='addition',
                        quantity=quantity,
                        purchase_price=purchase_price,
                        reason=notes or f'Stock incremented via Excel row {row_idx}',
                        created_by=created_by
                    )
                    stock_created = True
                else:
                    add_stock_to_warehouse(
                        product_id=product.id,
                        warehouse_id=main_wh.id,
                        quantity=quantity,
                        purchase_price=purchase_price,
                        selling_price=selling_price_final,
                        supplier_id=supplier.id if supplier else None,
                        notes=notes or f'Imported from Excel row {row_idx}',
                        created_by=created_by,
                    )
                    stock_created = True

            created_products.append({
                'product': product.name,
                'supplier': supplier.name if supplier else None,
                'row': row_idx,
                'notes': notes,
                'stock_created': stock_created,
            })
        except Exception as exc:
            errors.append({'row': row_idx, 'error': str(exc)})

    return {
        'created': created_products,
        'errors': errors,
        'summary': {
            'total_rows': ws.max_row - 1,
            'processed': len(created_products) + len(errors),
            'successful': len(created_products),
            'failed': len(errors),
        }
    }


def _get_user_branch(user):
    """Return the branch assigned to the logged-in user via Employee profile."""
    if user is None or not getattr(user, 'is_authenticated', False):
        return None
    employee = user.profile.first() if hasattr(user, 'profile') else None
    if employee and employee.branch:
        return employee.branch
    return None


def _get_selling_price_from_branch_stock(product: Product, branch) -> Optional[Decimal]:
    """Get the most recent selling price from branch stock for a product at a branch."""
    if branch is None:
        return None
    from stock.models import BranchStock
    branch_stock = BranchStock.objects.filter(
        product=product,
        branch=branch,
        quantity__gt=0,
    ).order_by('-received_date', '-created_at').first()
    if branch_stock and branch_stock.selling_price is not None:
        return branch_stock.selling_price
    return None


def _format_barcode_price_text(selling_price) -> Optional[str]:
    if selling_price is None:
        return None
    try:
        return f"USD ${Decimal(selling_price):.2f}"
    except Exception:
        return f"USD ${selling_price}"


def create_product(sku: str, name: str, category: int = None, unit: int = None, image=None, created_by=None) -> Product:
    """
    Create a new product and automatically generate a barcode linked to the product SKU.
    
    Args:
        sku: Product SKU
        name: Product name
        category: Category ID (optional)
        unit: Unit ID (optional)
        image: Product image file (optional)
    
    Returns:
        Created Product instance
    """
    # Get category and unit instances if IDs are provided
    category_instance = None
    if category is not None:
        try:
            category_instance = Category.objects.get(pk=category)
        except Category.DoesNotExist:
            category_instance = None
    
    unit_instance = None
    if unit is not None:
        try:
            unit_instance = Unit.objects.get(pk=unit)
        except Unit.DoesNotExist:
            unit_instance = None
    
    product = Product.objects.create(
        sku=sku,
        name=name,
        category=category_instance,
        unit=unit_instance,
        image=image
    )
    
    # Auto-generate barcode for the product using SKU as barcode value
    generate_product_barcode(product, user=created_by)
    
    return product


def generate_product_barcode(product: Product, user=None) -> 'Barcode':
    """
    Generate and create a barcode for a product using the product SKU.
    Uses CODE128 format (auto-generated).
    The barcode value will be the product SKU, making it easy to lookup products.
    
    Args:
        product: Product instance
        user: Logged-in user; selling price is taken from branch stock at the user's branch
        
    Returns:
        Created Barcode instance
    """
    from .models import Barcode
    from .utils import generate_barcode_image
    
    # Use product SKU as barcode value
    barcode_value = product.sku
    
    # Check if barcode already exists for this product
    existing_barcode = Barcode.objects.filter(product=product, barcode=barcode_value).first()
    if existing_barcode:
        if user is not None:
            regenerate_barcode_image(existing_barcode, user=user)
        return existing_barcode
    
    # Check if barcode value already exists for another product
    if Barcode.objects.filter(barcode=barcode_value).exclude(product=product).exists():
        # If SKU is already used, use product ID instead
        barcode_value = str(product.id)
    
    branch = _get_user_branch(user)
    selling_price = _get_selling_price_from_branch_stock(product, branch)
    price_text = _format_barcode_price_text(selling_price)

    # Generate barcode image (always CODE128) with price label when available
    barcode_image = generate_barcode_image(barcode_value, price_text=price_text)
    
    if not barcode_image:
        # If image generation fails, create barcode without image (frontend can add later)
        barcode = Barcode.objects.create(
            product=product,
            barcode=barcode_value,
            is_primary=True,
            notes=f'Auto-generated barcode for product {product.name}'
        )
        return barcode
    
    # Create barcode with generated image
    barcode = Barcode.objects.create(
        product=product,
        barcode_image=barcode_image,
        barcode=barcode_value,
        is_primary=True,
        notes=f'Auto-generated barcode for product {product.name}'
    )
    
    return barcode


def regenerate_barcode_image(barcode: Barcode, user=None) -> bool:
    """
    Regenerate barcode image with current selling price from the user's branch stock.
    
    Args:
        barcode: Barcode instance to regenerate
        user: Logged-in user; selling price is taken from branch stock at the user's branch
        
    Returns:
        True if regeneration was successful, False otherwise
    """
    from .utils import generate_barcode_image
    from django.core.files.storage import default_storage
    
    try:
        branch = _get_user_branch(user)
        selling_price = _get_selling_price_from_branch_stock(barcode.product, branch)
        price_text = _format_barcode_price_text(selling_price)
        
        # Generate new barcode image with current price
        barcode_image = generate_barcode_image(barcode.barcode, price_text=price_text)
        
        if not barcode_image:
            return False
        
        # Delete old image if it exists
        if barcode.barcode_image:
            if default_storage.exists(barcode.barcode_image.name):
                default_storage.delete(barcode.barcode_image.name)
        
        # Update barcode with new image
        barcode.barcode_image = barcode_image
        barcode.save()
        
        return True
        
    except Exception as e:
        print(f"Error regenerating barcode image for {barcode.barcode}: {str(e)}")
        return False


def add_product_to_main_warehouse(product: Product, warehouse_id: int, quantity: Decimal, 
                                  purchase_price: Decimal = None, supplier_id: int = None,
                                  batch_number: str = None, created_by=None):
    """
    Add product to main warehouse (initial stock).
    Now uses StockEntry system to track purchase prices.
    """
    # Create stock entry using the stock management system
    stock_entry = add_stock_to_warehouse(
        product_id=product.id,
        warehouse_id=warehouse_id,
        quantity=quantity,
        purchase_price=purchase_price,
        supplier_id=supplier_id,
        batch_number=batch_number,
        notes=f'Initial stock for product {product.name}',
        created_by=created_by
    )
    
    # Return the StockEntry (aggregated view)
    return stock_entry


def get_main_warehouse() -> Warehouse:
    """
    Get the main warehouse (is_main=True).
    Raises Warehouse.DoesNotExist if no main warehouse is set.
    """
    try:
        return Warehouse.objects.get(is_main=True)
    except Warehouse.DoesNotExist:
        raise Warehouse.DoesNotExist(
            'No main warehouse found. Please mark a warehouse as main (is_main=True) before creating products.'
        )
    except Warehouse.MultipleObjectsReturned:
        # If multiple main warehouses exist, get the first one
        return Warehouse.objects.filter(is_main=True).first()


def bulk_create_products(products_data: List[Dict], created_by=None) -> Tuple[List[Product], List[Dict]]:
    """
    Bulk create products with optional initial stock entries.
    
    Args:
        products_data: List of dictionaries containing product and optional stock information
            Each dict should have: sku, name, category (optional), unit (optional)
            Optional stock fields: initial_warehouse_id, initial_quantity, initial_purchase_price,
            batch_number, supplier_id. If initial_quantity is provided, stock will be created.
            If initial_warehouse_id is not provided and stock is being created, uses main warehouse.
        created_by: User creating the products
    
    Returns:
        Tuple of (created_products, errors)
        - created_products: List of successfully created Product objects
        - errors: List of error dictionaries with 'index' and 'error' keys
    """
    created_products = []
    errors = []
    
    with transaction.atomic():
        for index, product_data in enumerate(products_data):
            try:
                # Extract product fields
                sku = product_data.get('sku')
                name = product_data.get('name')
                category = product_data.get('category')
                unit = product_data.get('unit')
                image = product_data.get('image')  # Extract image if provided
                
                # Extract stock fields (optional)
                initial_warehouse_id = product_data.get('initial_warehouse_id')
                initial_quantity = product_data.get('initial_quantity')
                initial_purchase_price = product_data.get('initial_purchase_price')
                supplier_id = product_data.get('supplier_id')
                batch_number = product_data.get('batch_number')
                
                # Create product
                product = create_product(
                    sku=sku,
                    name=name,
                    category=category,
                    unit=unit,
                    image=image,
                    created_by=created_by,
                )
                
                # Only add initial stock if quantity is provided
                if initial_quantity is not None:
                    # If no warehouse specified, use main warehouse
                    if not initial_warehouse_id:
                        main_warehouse = get_main_warehouse()
                        initial_warehouse_id = main_warehouse.id
                    
                    # Add initial stock
                    add_product_to_main_warehouse(
                        product=product,
                        warehouse_id=initial_warehouse_id,
                        quantity=initial_quantity,
                        purchase_price=initial_purchase_price,
                        supplier_id=supplier_id,
                        batch_number=batch_number,
                        created_by=created_by
                    )
                
                created_products.append(product)
                
            except Exception as e:
                # Convert product_data to serializable format (convert objects to IDs)
                serializable_data = _make_serializable(product_data)
                
                errors.append({
                    'index': index,
                    'data': serializable_data,
                    'error': str(e)
                })
                # Continue with next product even if one fails
    
    return created_products, errors


def get_product_by_barcode(barcode_value: str) -> Optional[Product]:
    """
    Find a product by barcode value.
    Returns the product if found, None otherwise.
    
    Args:
        barcode_value: The barcode string to search for
        
    Returns:
        Product instance or None if not found
    """
    try:
        barcode = Barcode.objects.select_related('product').get(barcode=barcode_value)
        return barcode.product if barcode.product.is_active else None
    except Barcode.DoesNotExist:
        return None


def get_product_with_stock_by_barcode(barcode_value: str, branch_id: int = None, warehouse_id: int = None) -> Optional[Dict]:
    """
    Find a product by barcode and include available stock information.
    Links Product, Barcode, and Stock for frontend consumption.
    
    Args:
        barcode_value: The barcode string to search for
        branch_id: Optional branch ID to check stock availability at that branch
        warehouse_id: Optional warehouse ID to check stock availability at that warehouse
        
    Returns:
        Dictionary with product, barcode info, and stock info, or None if not found
        {
            'product': Product instance,
            'barcode': Barcode instance,
            'available_stock': Decimal (if branch_id or warehouse_id provided),
            'selling_price': Decimal (if branch_id provided and stock exists),
            'purchase_price': Decimal (if branch_id or warehouse_id provided and stock exists)
        }
    """
    try:
        barcode = Barcode.objects.select_related('product').get(barcode=barcode_value)
        product = barcode.product
        
        if not product.is_active:
            return None
        
        result = {
            'product': product,
            'barcode': barcode,
        }
        
        # If branch_id provided, get branch stock information
        if branch_id:
            from stock.models import BranchStock
            from django.db.models import Sum
            
            branch_stock = BranchStock.objects.filter(
                product=product,
                branch_id=branch_id,
                quantity__gt=0
            ).aggregate(total=Sum('quantity'))['total'] or Decimal('0')
            
            result['available_stock'] = branch_stock
            
            # Get selling price from most recent stock entry
            latest_stock = BranchStock.objects.filter(
                product=product,
                branch_id=branch_id,
                quantity__gt=0
            ).order_by('-received_date', '-created_at').first()
            
            if latest_stock:
                result['selling_price'] = latest_stock.selling_price
                result['purchase_price'] = latest_stock.purchase_price
            else:
                result['selling_price'] = None
                result['purchase_price'] = None
        
        # If warehouse_id provided, get warehouse stock information
        elif warehouse_id:
            from stock.models import StockEntry
            from django.db.models import Sum
            
            warehouse_stock = StockEntry.objects.filter(
                product=product,
                warehouse_id=warehouse_id,
                quantity__gt=0
            ).aggregate(total=Sum('quantity'))['total'] or Decimal('0')
            
            result['available_stock'] = warehouse_stock
            
            # Get purchase price from most recent stock entry
            latest_stock = StockEntry.objects.filter(
                product=product,
                warehouse_id=warehouse_id,
                quantity__gt=0
            ).order_by('-received_date', '-created_at').first()
            
            if latest_stock:
                result['purchase_price'] = latest_stock.purchase_price
            else:
                result['purchase_price'] = None
        
        return result
        
    except Barcode.DoesNotExist:
        return None


def create_barcode(product_id: int, barcode_value: str, barcode_image, 
                   is_primary: bool = False, notes: str = '') -> Barcode:
    """
    Create a Barcode instance with image and scanned value (provided by frontend).
    Uses CODE128 format (same as auto-generated barcodes).
    
    Args:
        product_id: ID of the product to link the barcode to
        barcode_value: The scanned barcode value (provided by frontend scanner)
        barcode_image: Django UploadedFile or file-like object containing barcode image
        is_primary: Whether this is the primary barcode for the product
        notes: Additional notes about the barcode
        
    Returns:
        Created Barcode instance
        
    Raises:
        ValueError: If barcode value already exists
    """
    product = Product.objects.get(pk=product_id)
    
    # Check if barcode already exists
    if Barcode.objects.filter(barcode=barcode_value).exists():
        raise ValueError(f"Barcode {barcode_value} already exists for another product.")
    
    # Create barcode instance (all barcodes use CODE128 format)
    barcode = Barcode.objects.create(
        product=product,
        barcode_image=barcode_image,
        barcode=barcode_value,
        is_primary=is_primary,
        notes=notes
    )
    
    return barcode